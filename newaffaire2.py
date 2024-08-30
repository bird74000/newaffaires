import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re
import os


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Formulaire de Saisie")
        self.file_name = "projet_data.xlsx"
        self.create_widgets()

    def validate_float_input(self, char):
        return char.isdigit() or char == "."

    def create_widgets(self):
        labels = [
            "Responsable Projet", "N° Devis", "N° d’Affaire", "Client", "DO",
            "Projet/Chantier", "Date de la Commande", "N° Commande", "Montant du Marché HT",
            "Observation", "Matière Prévue", "Sous-traitance Prévue",
            "Heure Chantier", "Heures Chantier 25%", "Étude", "Commentaire"
        ]

        self.entries = {}
        for i, label in enumerate(labels):
            tk.Label(self, text=label).grid(row=i, column=0, padx=10, pady=5, sticky=tk.W)
            entry = tk.Entry(self, width=40)
            if label in [
                "Montant du Marché HT", "Matière Prévue", "Sous-traitance Prévue",
                "Heure Chantier", "Heures Chantier 25%", "Étude"
            ]:
                validate_cmd = self.register(self.validate_float_input)
                entry.config(validate="key", validatecommand=(validate_cmd, "%S"))
            entry.grid(row=i, column=1, padx=10, pady=5)
            self.entries[label] = entry

        self.create_buttons(len(labels))

    def create_buttons(self, label_count):
        button_frame = tk.Frame(self)  # Crée un cadre pour contenir les boutons
        button_frame.grid(row=label_count, column=1, padx=10, pady=10, sticky="e")  # Place le cadre à droite

        tk.Button(button_frame, text="Valider", command=self.save_data, fg="blue", width=15).grid(row=0, column=0, pady=5, sticky="e")
        tk.Button(button_frame, text="Affaires", command=self.display_data, fg="green", width=15).grid(row=1, column=0, pady=5, sticky="e")
        tk.Button(button_frame, text="Nouvel État", command=self.nouvel_etat, fg="black", width=15).grid(row=2, column=0, pady=5, sticky="e")
        tk.Button(button_frame, text="Quitter", command=self.quit, fg="red", width=15).grid(row=3, column=0, pady=5, sticky="e")

        tk.Label(self, text="version 1 - YR", font=("Arial", 8)).grid(row=label_count + 2, column=1, pady=5, sticky="e")

    def load_affaires_en_cours(self):
        affaires = []
        if os.path.exists(self.file_name):
            workbook = load_workbook(self.file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if (row[-5] == "Oui" or row[-4] == "Oui" or row[-3] == "Oui") and row[-2] != "Oui":
                    affaires.append((row[2], row[3], row[5]))
            workbook.close()
        return affaires

    def nouvel_etat(self):
        def update_entries(*args):
            selected_affaire = affaire_var.get()
            for affaire in affaires_en_cours:
                if affaire[0] == selected_affaire:
                    client_entry.delete(0, tk.END)
                    client_entry.insert(0, affaire[1])
                    projet_entry.delete(0, tk.END)
                    projet_entry.insert(0, affaire[2])
                    break

        def save_etat():
            try:
                workbook = load_workbook(self.file_name)
                sheet = workbook.active
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    if row[2].value == affaire_var.get() and row[3].value == client_entry.get():
                        row[-5].value = "Oui" if litige_var.get() else "Non"
                        row[-4].value = "Oui" if en_cours_var.get() else "Non"
                        if facturer_var.get():
                            row[-3].value = "Oui"
                            row[-2].value = "Oui"
                        else:
                            row[-3].value = "Oui" if terminee_var.get() else "Non"
                            row[-2].value = "Non"
                        row[-1].value = commentaire_entry.get()
                        break
                workbook.save(self.file_name)
                workbook.close()
                messagebox.showinfo("Succès", "Le nouvel état a été sauvegardé avec succès.")
                new_window.destroy()
            except Exception as e:
                messagebox.showerror("Erreur", f"Erreur lors de la mise à jour de l'état : {e}")
            self.nouvel_etat()

        affaires_en_cours = self.load_affaires_en_cours()
        if not affaires_en_cours:
            messagebox.showinfo("Information", "Aucune affaire en cours ou terminée non facturée disponible.")
            return

        new_window = tk.Toplevel(self)
        new_window.title("Nouvel État")

        tk.Label(new_window, text="N° d’Affaire").grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
        affaire_var = tk.StringVar(new_window)
        affaire_var.set(affaires_en_cours[0][0])
        affaire_var.trace('w', update_entries)
        tk.OptionMenu(new_window, affaire_var, *[affaire[0] for affaire in affaires_en_cours]).grid(row=0, column=1, padx=10, pady=5, sticky="ew")

        tk.Label(new_window, text="Client").grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
        client_entry = tk.Entry(new_window, width=40)
        client_entry.grid(row=2, column=1, padx=10, pady=5)

        tk.Label(new_window, text="Projet/Chantier").grid(row=3, column=0, padx=10, pady=5, sticky=tk.W)
        projet_entry = tk.Entry(new_window, width=40)
        projet_entry.grid(row=3, column=1, padx=10, pady=5)

        update_entries()

        en_cours_var = tk.BooleanVar()
        tk.Checkbutton(new_window, text="En cours", variable=en_cours_var).grid(row=4, column=1, padx=10, pady=5, sticky=tk.W)

        terminee_var = tk.BooleanVar()
        tk.Checkbutton(new_window, text="Terminé", variable=terminee_var).grid(row=5, column=1, padx=10, pady=5, sticky=tk.W)

        facturer_var = tk.BooleanVar()
        tk.Checkbutton(new_window, text="Facturé", variable=facturer_var).grid(row=6, column=1, padx=10, pady=5, sticky=tk.W)

        litige_var = tk.BooleanVar()
        tk.Checkbutton(new_window, text="Litige", variable=litige_var).grid(row=7, column=1, padx=10, pady=5, sticky=tk.W)

        tk.Label(new_window, text="Commentaire").grid(row=8, column=0, padx=10, pady=5, sticky=tk.W)
        commentaire_entry = tk.Entry(new_window, width=40)
        commentaire_entry.grid(row=8, column=1, padx=10, pady=5)

        tk.Button(new_window, text="Sauvegarder", command=save_etat).grid(row=9, column=1, padx=10, pady=10, sticky="ew")
        tk.Button(new_window, text="Fermer", command=new_window.destroy).grid(row=9, column=0, padx=10, pady=10, sticky="ew")

    def save_data(self):
        try:
            data = self.validate_and_collect_data()
        except ValueError as e:
            messagebox.showerror("Erreur de validation", str(e))
            return

        if os.path.exists(self.file_name):
            workbook = load_workbook(self.file_name)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            for i, key in enumerate(data.keys(), 1):
                sheet[f'{get_column_letter(i)}1'] = key

        row = sheet.max_row + 1
        for i, value in enumerate(data.values(), 1):
            sheet[f'{get_column_letter(i)}{row}'] = value

        workbook.save(self.file_name)
        workbook.close()
        self.clear_form()
        messagebox.showinfo("Succès", "Les données ont été sauvegardées avec succès.")

    def validate_and_collect_data(self):
        date_pattern = r"^\d{2}/\d{2}/\d{4}$"
        float_fields = ["Montant du Marché HT", "Matière Prévue", "Sous-traitance Prévue", "Heure Chantier", "Heures Chantier 25%", "Étude"]

        if not re.match(date_pattern, self.entries["Date de la Commande"].get()):
            raise ValueError("La date de commande doit être au format JJ/MM/AAAA.")

        data = {}
        for key, entry in self.entries.items():
            if key in float_fields:
                try:
                    data[key] = float(entry.get())
                except ValueError:
                    raise ValueError(f"Le champ '{key}' doit contenir un nombre valide.")
            else:
                data[key] = entry.get()

        data.update({
            "Litige": "Non",
            "En cours": "Oui",
            "Terminé": "Non",
            "Facturé": "Non",
            "Commentaire": ""
        })

        return data

    def clear_form(self):
        for entry in self.entries.values():
            entry.delete(0, tk.END)

    def display_data(self):
        if not os.path.exists(self.file_name):
            messagebox.showwarning("Avertissement", "Le fichier de données n'existe pas encore.")
            return

        workbook = load_workbook(self.file_name)
        sheet = workbook.active
        display_window = tk.Toplevel(self)
        display_window.title("Affichage des données")
        tree = ttk.Treeview(display_window, columns=list(range(1, sheet.max_column + 1)), show='headings')

        for col_num, col_name in enumerate(sheet.iter_cols(1, sheet.max_column, 1, 1, True), start=1):
            tree.heading(col_num, text=col_name[0])
            tree.column(col_num, width=100, anchor='center')

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            item = tree.insert('', tk.END, values=row)
            if row[-5] == "Oui":  # Si litige est "Oui"
                tree.item(item, tags=("litige",))
                tree.tag_configure("litige", background="red")
            elif row[-3] == "Oui" and row[-2] == "Oui":
                tree.item(item, tags=("invoiced",))
                tree.tag_configure("invoiced", background="lightgreen")
            elif row[-3] == "Oui":
                tree.item(item, tags=("completed",))
                tree.tag_configure("completed", background="orange")
            elif row[-4] == "Oui":
                tree.item(item, tags=("in_progress",))
                tree.tag_configure("in_progress", background="yellow")

        tree.pack(expand=True, fill='both')

        # Bind the double-click event to the Treeview
        tree.bind("<Double-1>", self.show_details)

        display_window.mainloop()

    def show_details(self, event):
        # Get the selected item
        tree = event.widget
        selected_item = tree.selection()[0]
        values = tree.item(selected_item, "values")

        # Create a new window to display the details
        detail_window = tk.Toplevel(self)
        detail_window.title("Détails de l'affaire")
        detail_window.geometry("500x400")

        # Create a canvas to hold the details and add a scrollbar
        canvas = tk.Canvas(detail_window)
        scrollbar = tk.Scrollbar(detail_window, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(
                scrollregion=canvas.bbox("all")
            )
        )

        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        labels = [
            "Responsable Projet", "N° Devis", "N° d’Affaire", "Client", "DO",
            "Projet/Chantier", "Date de la Commande", "N° Commande", "Montant du Marché HT",
            "Observation", "Matière Prévue", "Sous-traitance Prévue",
            "Heure Chantier", "Heures Chantier 25%", "Étude", "Litige", "En cours",
            "Terminé", "Facturé", "Commentaire"
        ]

        for i, (label, value) in enumerate(zip(labels, values)):
            tk.Label(scrollable_frame, text=f"{label} :").grid(row=i, column=0, padx=10, pady=5, sticky=tk.W)
            tk.Label(scrollable_frame, text=value).grid(row=i, column=1, padx=10, pady=5, sticky=tk.W)

        tk.Button(scrollable_frame, text="Fermer", command=detail_window.destroy).grid(row=len(labels), column=1, padx=10, pady=10, sticky="ew")

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")


if __name__ == "__main__":
    app = Application()
    app.mainloop()
