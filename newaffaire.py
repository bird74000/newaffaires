import tkinter as tk
from tkinter import messagebox, ttk
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import re
import os

file_name = "projet_data.xlsx"

# Fonction pour permettre seulement les nombres dans certains champs
def validate_float_input(char):
    return char.isdigit() or char == "."

# Fonction pour charger les affaires en cours ou terminées mais non facturées depuis le fichier Excel
def load_affaires_en_cours():
    affaires = []
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if (row[-3] == "Oui" or row[-2] == "Oui") and row[-1] != "Oui":  # Affaire en cours ou terminée mais non facturée
                affaires.append((row[2], row[3], row[5]))  # (N° d'Affaire, Client, Projet/Chantier)
        workbook.close()
    return affaires

# Fonction pour afficher la boîte de dialogue du nouvel état
def nouvel_etat():
    def update_entries(*args):
        selected_affaire = affaire_var.get()
        for affaire in affaires_en_cours:
            if affaire[0] == selected_affaire:
                client_entry.delete(0, tk.END)
                client_entry.insert(0, affaire[1])
                projet_entry.delete(0, tk.END)
                projet_entry.insert(0, affaire[2])
                break

    def save_etat():
        # Mettre à jour l'état dans le fichier Excel
        try:
            workbook = load_workbook(file_name)
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                if row[2].value == affaire_var.get() and row[3].value == client_entry.get():
                    row[-3].value = "Oui" if en_cours_var.get() else "Non"
                    # Si "Facturé" est "Oui", "Terminé" doit également être "Oui"
                    if facturer_var.get():
                        row[-2].value = "Oui"
                        row[-1].value = "Oui"
                    else:
                        row[-2].value = "Oui" if terminee_var.get() else "Non"
                        row[-1].value = "Non"
                    break
            workbook.save(file_name)
            workbook.close()
            messagebox.showinfo("Succès", "Le nouvel état a été sauvegardé avec succès.")
            new_window.destroy()
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de la mise à jour de l'état : {e}")
        nouvel_etat()  # Réouvrir la fenêtre pour actualiser la liste

    affaires_en_cours = load_affaires_en_cours()
    if not affaires_en_cours:
        messagebox.showinfo("Information", "Aucune affaire en cours ou terminée non facturée disponible.")
        return

    new_window = tk.Toplevel(root)
    new_window.title("Nouvel État")

    tk.Label(new_window, text="N° d’Affaire").grid(row=0, column=0, padx=10, pady=5, sticky=tk.W)
    affaire_var = tk.StringVar(new_window)
    affaire_var.set(affaires_en_cours[0][0])
    affaire_var.trace('w', update_entries)
    affaire_menu = tk.OptionMenu(new_window, affaire_var, *[affaire[0] for affaire in affaires_en_cours])
    affaire_menu.grid(row=0, column=1, padx=10, pady=5, sticky="ew")

    tk.Label(new_window, text="Client").grid(row=2, column=0, padx=10, pady=5, sticky=tk.W)
    client_entry = tk.Entry(new_window, width=40)
    client_entry.grid(row=2, column=1, padx=10, pady=5)

    tk.Label(new_window, text="Projet/Chantier").grid(row=3, column=0, padx=10, pady=5, sticky=tk.W)
    projet_entry = tk.Entry(new_window, width=40)
    projet_entry.grid(row=3, column=1, padx=10, pady=5)

    # Mettre à jour les champs "Client" et "Projet/Chantier" lors de l'ouverture du formulaire
    update_entries()

    en_cours_var = tk.BooleanVar()
    en_cours_check = tk.Checkbutton(new_window, text="En cours", variable=en_cours_var)
    en_cours_check.grid(row=4, column=1, padx=10, pady=5, sticky=tk.W)

    terminee_var = tk.BooleanVar()
    terminee_check = tk.Checkbutton(new_window, text="Terminé", variable=terminee_var)
    terminee_check.grid(row=5, column=1, padx=10, pady=5, sticky=tk.W)

    facturer_var = tk.BooleanVar()
    facturer_check = tk.Checkbutton(new_window, text="Facturé", variable=facturer_var)
    facturer_check.grid(row=6, column=1, padx=10, pady=5, sticky=tk.W)

    tk.Button(new_window, text="Sauvegarder", command=save_etat).grid(row=7, column=1, padx=10, pady=10, sticky="ew")
    tk.Button(new_window, text="Fermer", command=new_window.destroy).grid(row=7, column=0, padx=10, pady=10, sticky="ew")

# Fonction pour valider et sauvegarder les données dans un fichier Excel
def save_data():
    date_pattern = r"^\d{2}/\d{2}/\d{4}$"
    float_pattern = r"^\d+(\.\d{1,2})?$"

    # Validation des champs
    if not re.match(date_pattern, entry_date_commande.get()):
        messagebox.showerror("Erreur de validation", "La date de commande doit être au format JJ/MM/AAAA.")
        return
    
    # Conversion et validation des champs numériques
    try:
        montant_marche_ht = float(entry_montant.get())
        matiere_prevue = float(entry_matiere.get())
        sous_traitance_prevue = float(entry_sous_traitance.get())
        heure_chantier = float(entry_heure_chantier.get())
        heures_chantier_25 = float(entry_heures_25.get())
        etude = float(entry_etude.get())
    except ValueError:
        messagebox.showerror("Erreur de validation", "Tous les champs numériques doivent contenir des nombres valides.")
        return

    data = {
        "Responsable Projet": entry_responsable.get(),
        "N° Devis": entry_devis.get(),
        "N° d’Affaire": entry_affaire.get(),
        "Client": entry_client.get(),
        "DO": entry_do.get(),
        "Projet/Chantier": entry_projet.get(),
        "Date de la Commande": entry_date_commande.get(),
        "N° Commande": entry_commande.get(),
        "Montant du Marché HT": montant_marche_ht,
        "Observation": entry_observation.get(),
        "Matière Prévue": matiere_prevue,
        "Sous-traitance Prévue": sous_traitance_prevue,
        "Heure Chantier": heure_chantier,
        "Heures Chantier 25%": heures_chantier_25,
        "Étude": etude,
        "En cours": "Oui",  # Par défaut, marquer comme "En cours"
        "Terminé": "Non",
        "Facturé": "Non"
    }

    # Vérifier si le fichier Excel existe déjà
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Créer les en-têtes
        for i, key in enumerate(data.keys(), 1):
            sheet[f'{get_column_letter(i)}1'] = key

    # Ajouter les données à la fin de la feuille
    row = sheet.max_row + 1
    for i, value in enumerate(data.values(), 1):
        sheet[f'{get_column_letter(i)}{row}'] = value

    workbook.save(file_name)
    workbook.close()

    # Réinitialiser le formulaire après la sauvegarde
    clear_form()

    messagebox.showinfo("Succès", "Les données ont été sauvegardées avec succès.")

# Fonction pour réinitialiser le formulaire
def clear_form():
    entry_responsable.delete(0, tk.END)
    entry_devis.delete(0, tk.END)
    entry_affaire.delete(0, tk.END)
    entry_client.delete(0, tk.END)
    entry_do.delete(0, tk.END)
    entry_projet.delete(0, tk.END)
    entry_date_commande.delete(0, tk.END)
    entry_commande.delete(0, tk.END)
    entry_montant.delete(0, tk.END)
    entry_observation.delete(0, tk.END)
    entry_matiere.delete(0, tk.END)
    entry_sous_traitance.delete(0, tk.END)
    entry_heure_chantier.delete(0, tk.END)
    entry_heures_25.delete(0, tk.END)
    entry_etude.delete(0, tk.END)

# Fonction pour afficher les données dans une nouvelle fenêtre
def display_data():
    global workbook, file_name
    
    if not os.path.exists(file_name):
        messagebox.showwarning("Avertissement", "Le fichier de données n'existe pas encore.")
        return

    workbook = load_workbook(file_name)
    sheet = workbook.active
    
    display_window = tk.Toplevel(root)
    display_window.title("Affichage des données")
    
    tree = ttk.Treeview(display_window, columns=list(range(1, sheet.max_column + 1)), show='headings')
    
    for col_num, col_name in enumerate(sheet.iter_cols(1, sheet.max_column, 1, 1, True), start=1):
        tree.heading(col_num, text=col_name[0])
        tree.column(col_num, width=100, anchor='center')
    
    for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True), start=2):
        values = list(row)
        item = tree.insert('', tk.END, values=values)
        
        if values[-2] == "Oui" and values[-1] == "Oui":
            tree.item(item, tags=("invoiced",))
            tree.tag_configure("invoiced", background="lightgreen")
        elif values[-2] == "Oui":
            tree.item(item, tags=("completed",))
            tree.tag_configure("completed", background="orange")
        elif values[-3] == "Oui":
            tree.item(item, tags=("in_progress",))
            tree.tag_configure("in_progress", background="yellow")
    
    tree.pack(expand=True, fill='both')
    display_window.mainloop()

# Fonction pour quitter l'application
def quitter_application():
    root.quit()

# Créer l'interface utilisateur
root = tk.Tk()
root.title("Formulaire de Saisie")

labels = ["Responsable Projet", "N° Devis", "N° d’Affaire", "Client", "DO", 
          "Projet/Chantier", "Date de la Commande", "N° Commande", "Montant du Marché HT", 
          "Observation", "Matière Prévue", "Sous-traitance Prévue", 
          "Heure Chantier", "Heures Chantier 25%", "Étude"]

entries = []

for i, label in enumerate(labels):
    tk.Label(root, text=label).grid(row=i, column=0, padx=10, pady=5, sticky=tk.W)
    entry = tk.Entry(root, width=40)
    if label in ["Montant du Marché HT", "Matière Prévue", "Sous-traitance Prévue", 
                 "Heure Chantier", "Heures Chantier 25%", "Étude"]:
        validate_cmd = root.register(validate_float_input)
        entry.config(validate="key", validatecommand=(validate_cmd, "%S"))
    entry.grid(row=i, column=1, padx=10, pady=5)
    entries.append(entry)

(entry_responsable, entry_devis, entry_affaire, entry_client, entry_do, 
 entry_projet, entry_date_commande, entry_commande, entry_montant, 
 entry_observation, entry_matiere, entry_sous_traitance, 
 entry_heure_chantier, entry_heures_25, entry_etude) = entries

save_button = tk.Button(root, text="Valider", command=save_data, fg="blue", width=15)
save_button.grid(row=len(labels), column=0, pady=10, sticky="ew")

display_button = tk.Button(root, text="Affaires", command=display_data, fg="green", width=15)
display_button.grid(row=len(labels), column=1, pady=10, sticky="ew")

nouvel_etat_button = tk.Button(root, text="Nouvel État", command=nouvel_etat, fg="black", width=15)
nouvel_etat_button.grid(row=len(labels), column=2, pady=10, sticky="ew")

quit_button = tk.Button(root, text="Quitter", command=quitter_application, fg="red", width=15)
quit_button.grid(row=len(labels)+1, column=1, pady=10, sticky="ew")

# Ajouter le label de version en bas du formulaire
version_label = tk.Label(root, text="version 1 - YR", font=("Arial", 8))
version_label.grid(row=len(labels)+2, column=1, pady=5, sticky="e")

root.mainloop()
